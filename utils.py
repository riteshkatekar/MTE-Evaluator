# utils.py
import json
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles.borders import Border
from collections import deque
import os
import re
from datetime import datetime
from openpyxl.styles import Border

def load_json():
    """
    Load JSON data from the file and return it as a dictionary.
    """
    try:
        with open("system_config.json", 'r') as file:
            data = json.load(file)
        return data
    except Exception as e:
        print(f"Error loading JSON file: {e}")
        return None

def get_api_key_from_json(key_name):
    """
    Get the API key from a JSON file.
    """
    config_data = load_json()
    if config_data and key_name in config_data:
        return config_data[key_name]
    else:
        print(f"{key_name} not found in the JSON file.")
        return None


def extract_mte_data(file_path):
    def normalize(value):
        if not isinstance(value, str):
            return ""
        return unicodedata.normalize("NFKD", value).strip()

    def has_border(cell):
        border: Border = cell.border
        sides = [border.left, border.right, border.top, border.bottom]
        return any(side.style is not None for side in sides)

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.worksheets[0]

        # --- Extract raw metadata lines (Rows 2 and 4 expected, fallback if missing) ---
        raw_student_info = normalize(sheet.cell(2, 2).value or "")
        raw_college_info = normalize(sheet.cell(4, 2).value or "")

        # --- Extract Name and Month robustly ---
        student_name, submission_month = "N/A", "N/A"
        student_match = re.search(r"Student\s*[:\-]?\s*(.+?)\s+for\s+the\s+month\s+of\s+([A-Za-z]+)", raw_student_info, re.IGNORECASE)
        if student_match:
            student_name = student_match.group(1).strip()
            submission_month = f"{student_match.group(2).capitalize()} , {datetime.now().year}"

        # --- Clean and Normalize College/Class Line ---
        raw_college_info_clean = re.sub(r"[_\s]+", " ", raw_college_info).strip()

        college_name, class_info = "N/A", "N/A"
        match_college = re.search(r"College\s*[:\-]?\s*(.+?)\s+Year\s+of\s+Study", raw_college_info_clean, re.IGNORECASE)
        match_year = re.search(r"Year\s+of\s+Study\s*[:\-]?\s*(.+)", raw_college_info_clean, re.IGNORECASE)
        if match_college:
            college_name = match_college.group(1).strip()
        if match_year:
            class_info = match_year.group(1).strip()

        # --- Extract bordered tables ---
        border_map = {}
        for row in sheet.iter_rows():
            for cell in row:
                if has_border(cell):
                    border_map[(cell.row, cell.column)] = normalize(cell.value)

        visited = set()
        groups = []
        directions = [(-1, 0), (1, 0), (0, -1), (0, 1)]

        for cell in border_map:
            if cell not in visited:
                group = []
                queue = deque([cell])
                visited.add(cell)
                while queue:
                    current = queue.popleft()
                    group.append(current)
                    for dr, dc in directions:
                        neighbor = (current[0] + dr, current[1] + dc)
                        if neighbor in border_map and neighbor not in visited:
                            visited.add(neighbor)
                            queue.append(neighbor)
                groups.append(group)

        extracted_tables = []
        for group in groups:
            rows = [r for r, _ in group]
            cols = [c for _, c in group]
            min_row, max_row = min(rows), max(rows)
            min_col, max_col = min(cols), max(cols)

            table = []
            for r in range(min_row, max_row + 1):
                row_data = []
                for c in range(min_col, max_col + 1):
                    value = normalize(sheet.cell(r, c).value)
                    if value:
                        row_data.append(value)
                if row_data:
                    table.append(row_data)

            if table:
                heading = " ".join(table[0])
                rows_as_string = "\n".join("    " + " | ".join(row) for row in table[1:])
                extracted_tables.append({"heading": heading.strip(), "rows": rows_as_string.strip()})

        # --- Map headings to dictionary keys ---
        mte_dict = {}
        heading_to_key_map = {
            "Academic Progress / Vacation Plan": "academic_progress",
            "Co and Extra Curricular Progress-Plan": "co-curricular",
            "Fin Reqm for the next 3 months (Details Please)": "financial_needs",
            "Difficulties (Social, Family, etc.)": "difficulties",
            "Results of the exams": "exam_results",
            "Reading Books / Watching Videos": "books_and_videos",
            "exercise regularly and eat and sleep": "health",
            "friends or acquaintances made": "learning_from_people",
            "Essay on a topic of your choice": "essay",
            "Action Plan for the coming month": "action_plan"
        }

        for table in extracted_tables:
            for expected_heading, dict_key in heading_to_key_map.items():
                if expected_heading.lower() in table['heading'].lower():
                    mte_dict[dict_key] = table['rows']
                    break

        return {
            "student_name": student_name,
            "submission_month": submission_month,
            "college_name": college_name,
            "class_info": class_info,
            "academic_progress": mte_dict.get("academic_progress", ""),
            "co-curricular": mte_dict.get("co-curricular", ""),
            "financial_needs": mte_dict.get("financial_needs", ""),
            "difficulties": mte_dict.get("difficulties", ""),
            "exam_results": mte_dict.get("exam_results", ""),
            "books_and_videos": mte_dict.get("books_and_videos", ""),
            "health": mte_dict.get("health", ""),
            "learning_from_people": mte_dict.get("learning_from_people", ""),
            "essay": mte_dict.get("essay", ""),
            "action_plan": mte_dict.get("action_plan", "")
        }

    except Exception as e:
        return {"error": f"Error extracting data: {e}"}



def save_feedback_to_json(feedback, student_id):
    """
    Saves the generated feedback to a JSON file to track historical data for students.
    """
    try:
        history_dir = "../data/history"
        file_path = f"{history_dir}/{student_id}_feedback.json"
        
        os.makedirs(history_dir, exist_ok=True)
        
        with open(file_path, "w") as f:
            json.dump(feedback, f, indent=4)
        
        return True
    except Exception as e:
        return {"error": f"Error saving feedback: {e}"}
